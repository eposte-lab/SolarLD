"""B2C door-to-door export endpoints.

Field agents working a territory on foot need a printable brief with
the CAP, income bucket, estimated household count, and a few lines of
suggested talk-track. We don't have civic-level addresses in the B2C
pipeline (everything is CAP-aggregated from ISTAT + census), so the
export is not a door list — it's a **CAP dossier** per audience that
an agent carries into the field.

Two formats:

  GET /v1/b2c/audiences/{id}/export.pdf
      → reportlab single-page A4 dossier: header with brand, stats
        block, talk-track bullets, map placeholder.

  GET /v1/b2c/audiences/{id}/export.xlsx
      → openpyxl workbook: ``Dossier`` sheet (same fields as PDF) +
        ``Contatti`` sheet (blank template the agent fills in as they
        knock — name, civic, exterior signals, outcome).

Both endpoints hit the same data-loader so swapping the template
between PDF and xlsx doesn't require keeping two queries in sync.

GDPR note: we deliberately do not include names / addresses / phone
numbers in the export. Only CAP-level statistics and blank fields the
agent populates by hand. No personal data leaves the system via this
route.
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import Response

from ..core.logging import get_logger
from ..core.security import CurrentUser, require_tenant
from ..core.supabase_client import get_service_client
from ..services.b2c_audience_service import get_audience
from ..services.tenant_module_service import get_module

router = APIRouter()
log = get_logger(__name__)


# ---------------------------------------------------------------------------
# Shared loader
# ---------------------------------------------------------------------------


async def _load_dossier_data(
    audience_id: UUID, tenant_id: UUID
) -> dict[str, Any]:
    """Resolve the audience + its ISTAT stats + outreach copy into one
    dict. Raises 404 if the audience doesn't exist for the tenant.
    """
    audience = await get_audience(audience_id, tenant_id)
    if not audience:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="audience not found",
        )

    sb = get_service_client()
    cap = audience.get("cap")
    geo_res = (
        sb.table("geo_income_stats")
        .select(
            "cap, provincia, regione, comune, reddito_medio_eur, "
            "popolazione, case_unifamiliari_pct"
        )
        .eq("cap", cap)
        .limit(1)
        .execute()
    )
    geo_rows = getattr(geo_res, "data", None) or []
    geo = geo_rows[0] if geo_rows else {}

    outreach_mod = await get_module(tenant_id, "outreach")
    ocfg = outreach_mod.config or {}

    return {
        "audience": audience,
        "geo": geo,
        "cta_primary": ocfg.get("cta_primary")
        or "Prenota un sopralluogo gratuito",
        "tone": ocfg.get("tone") or "cordiale, competente, non invasivo",
        "generated_on": date.today().isoformat(),
    }


def _talk_track(data: dict[str, Any]) -> list[str]:
    """Stock talk-track bullets the agent can skim before each knock.

    Kept intentionally generic — installers personalise via the
    outreach module's `tone` + `cta_primary`, but the underlying
    script structure is the same everywhere (greet → local hook →
    value prop → CTA)."""
    a = data["audience"]
    g = data["geo"]
    comune = g.get("comune") or a.get("provincia") or "la zona"
    bucket_label = {
        "basso": "reddito medio basso",
        "medio": "reddito medio",
        "alto": "reddito medio-alto",
        "premium": "reddito alto",
    }.get(a.get("reddito_bucket", "medio"), "reddito medio")

    return [
        f"Buongiorno, siamo in giro oggi per {comune} — "
        f"stiamo raccogliendo adesioni per un sopralluogo fotovoltaico "
        f"gratuito riservato a famiglie del vostro CAP.",
        f"Nel CAP {a.get('cap')} stiamo vedendo molte case unifamiliari "
        f"con tetto orientato bene — proprio il tipo di casa dove il "
        f"fotovoltaico rende di più.",
        "Un impianto tipico per una famiglia qui recupera il costo in "
        "circa 5-7 anni, poi produce energia gratis per altri 20.",
        f"Contesto zona: {bucket_label}. Target conversazionale: "
        f"{data['tone']}.",
        f"Chiusura: {data['cta_primary']}.",
    ]


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


@router.get("/audiences/{audience_id}/export.pdf")
async def export_audience_pdf(
    ctx: CurrentUser, audience_id: UUID
) -> Response:
    """A4 dossier PDF for one audience.

    Generated in memory with reportlab — we don't persist the file;
    the dashboard downloads a fresh copy each click so the talk-track
    always reflects the latest outreach module config.
    """
    tenant_id = require_tenant(ctx)
    data = await _load_dossier_data(audience_id, tenant_id)

    # Lazy imports: reportlab is ~40MB of class hierarchy and we don't
    # want to pay the import cost on every app boot just for this
    # occasional endpoint.
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Dossier CAP {data['audience'].get('cap')}",
    )

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle(
        "H", parent=styles["Heading1"], spaceAfter=6 * mm, fontSize=18
    )
    body_style = styles["BodyText"]
    small = ParagraphStyle(
        "Small", parent=body_style, fontSize=9, textColor=colors.grey
    )

    a = data["audience"]
    g = data["geo"]
    story: list[Any] = []

    story.append(
        Paragraph(
            f"Dossier Door-to-Door — CAP {a.get('cap')}", h_style
        )
    )
    story.append(
        Paragraph(
            f"{g.get('comune') or '—'} ({g.get('provincia') or '—'}) · "
            f"{g.get('regione') or '—'} · generato il "
            f"{data['generated_on']}",
            small,
        )
    )
    story.append(Spacer(1, 6 * mm))

    stats_rows = [
        ["Reddito medio", f"€{g.get('reddito_medio_eur') or '—'}"],
        ["Bucket", a.get("reddito_bucket", "—")],
        ["Popolazione CAP", f"{g.get('popolazione') or '—'}"],
        [
            "Case unifamiliari",
            f"{g.get('case_unifamiliari_pct') or '—'}%",
        ],
        [
            "Stima contatti",
            f"{a.get('stima_contatti') or '—'} famiglie",
        ],
    ]
    tbl = Table(stats_rows, colWidths=[55 * mm, 110 * mm])
    tbl.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("<b>Talk-track</b>", body_style))
    for line in _talk_track(data):
        story.append(Paragraph(f"• {line}", body_style))
    story.append(Spacer(1, 6 * mm))

    story.append(
        Paragraph(
            "<b>Note agente</b> (compilare a mano durante il giro)",
            body_style,
        )
    )
    # Empty lined block — 15 rows × empty cells — gives the agent a
    # writable surface for on-site notes without leaving the dossier.
    notes_rows = [[""] for _ in range(15)]
    notes_tbl = Table(notes_rows, colWidths=[165 * mm], rowHeights=[8 * mm] * 15)
    notes_tbl.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
                (
                    "LINEBELOW",
                    (0, 0),
                    (-1, -1),
                    0.25,
                    colors.lightgrey,
                ),
            ]
        )
    )
    story.append(notes_tbl)

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()

    log.info(
        "b2c_exports.pdf",
        extra={
            "tenant_id": str(tenant_id),
            "audience_id": str(audience_id),
            "cap": a.get("cap"),
            "bytes": len(pdf_bytes),
        },
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="dossier_CAP_{a.get("cap")}.pdf"'
            )
        },
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


@router.get("/audiences/{audience_id}/export.xlsx")
async def export_audience_xlsx(
    ctx: CurrentUser, audience_id: UUID
) -> Response:
    """Workbook export — ``Dossier`` + blank ``Contatti`` sheet.

    The Contatti sheet is intentionally empty — it's the form the
    agent fills in by hand (digitally) after the round. We don't
    pre-fill addresses because we don't have them (CAP-level only,
    GDPR-safe).
    """
    tenant_id = require_tenant(ctx)
    data = await _load_dossier_data(audience_id, tenant_id)

    # Lazy import — same reasoning as reportlab above.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # --- Sheet 1: Dossier ------------------------------------------------
    ws = wb.active
    ws.title = "Dossier"

    a = data["audience"]
    g = data["geo"]

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    header_fill = PatternFill(
        "solid", fgColor="EEF2FF"
    )  # pale indigo — matches dashboard brand

    ws["A1"] = f"Dossier Door-to-Door — CAP {a.get('cap')}"
    ws["A1"].font = header_font
    ws.merge_cells("A1:B1")

    ws["A2"] = (
        f"{g.get('comune') or '—'} ({g.get('provincia') or '—'}) · "
        f"{g.get('regione') or '—'}"
    )
    ws.merge_cells("A2:B2")
    ws["A3"] = f"Generato: {data['generated_on']}"
    ws.merge_cells("A3:B3")

    stats = [
        ("Reddito medio", f"€{g.get('reddito_medio_eur') or '—'}"),
        ("Bucket", a.get("reddito_bucket", "—")),
        ("Popolazione CAP", g.get("popolazione") or "—"),
        (
            "Case unifamiliari",
            f"{g.get('case_unifamiliari_pct') or '—'}%",
        ),
        ("Stima contatti", a.get("stima_contatti") or "—"),
    ]
    for i, (label, value) in enumerate(stats, start=5):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=1).fill = header_fill
        ws.cell(row=i, column=2, value=value)

    ws.cell(row=11, column=1, value="Talk-track").font = header_font
    for i, line in enumerate(_talk_track(data), start=12):
        c = ws.cell(row=i, column=1, value=f"• {line}")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 42
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    # --- Sheet 2: Contatti (blank) --------------------------------------
    ws2 = wb.create_sheet("Contatti")
    contatti_headers = [
        "Data visita",
        "Via + civico",
        "Citofono",
        "Tipologia casa",
        "Orientamento tetto",
        "Ombreggiamento",
        "Nome contatto",
        "Esito",
        "Prossimo step",
        "Note",
    ]
    for col_idx, h in enumerate(contatti_headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = label_font
        c.fill = header_fill
    # Sensible widths — each column tuned for its content.
    widths = [12, 30, 12, 18, 18, 14, 22, 14, 22, 40]
    for col_idx, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(ord("A") + col_idx - 1)].width = w

    # Pre-size for a day's work: 60 blank rows the agent can fill.
    for r in range(2, 62):
        ws2.row_dimensions[r].height = 18

    # Serialise
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    buf.close()

    log.info(
        "b2c_exports.xlsx",
        extra={
            "tenant_id": str(tenant_id),
            "audience_id": str(audience_id),
            "cap": a.get("cap"),
            "bytes": len(xlsx_bytes),
        },
    )
    return Response(
        content=xlsx_bytes,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="dossier_CAP_{a.get("cap")}.xlsx"'
            )
        },
    )
