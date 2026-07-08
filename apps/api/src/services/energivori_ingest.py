"""Fase 0 — ingest the CSEA "energivori" list (XLSX) into normalised records.

The energy-intensive companies list (≈5.7k rows) is the input of the new
OpenAPI discovery channel. Columns (sheet "Energivori 2026"):
``N., P.IVA, Codice Fiscale, Ragione Sociale, Settore (stima), Priorita FV
(stima), Classe, TipoClasse, Decorrenza, Stato``.

The parsing CORE (``parse_energivori_rows``) is PURE — it takes the header +
data rows, so it is fully unit-testable without touching a file. The thin
``parse_energivori_xlsx`` wrapper reads the workbook (lazy openpyxl import).

Critical detail: an Italian P.IVA is an 11-DIGIT string; leading zeros MUST be
preserved. If a source hands the value as an int (zeros stripped) we zero-pad
back to 11. Rows without a valid 11-digit P.IVA are skipped; duplicates are
collapsed on the P.IVA.
"""

from __future__ import annotations

import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from typing import Any

_DIGITS = re.compile(r"\D+")

# Header label (normalised: lower, stripped) → record field.
_COLUMN_MAP = {
    "p.iva": "piva",
    "piva": "piva",
    "partita iva": "piva",
    "codice fiscale": "codice_fiscale",
    "ragione sociale": "ragione_sociale",
    "settore (stima)": "settore",
    "settore": "settore",
    "priorita fv (stima)": "priorita_fv",
    "priorità fv (stima)": "priorita_fv",
    "classe": "classe",
    "tipoclasse": "tipo_classe",
    "stato": "stato",
}


@dataclass(frozen=True)
class EnergivoroRecord:
    piva: str  # 11-digit, leading zeros preserved
    ragione_sociale: str
    codice_fiscale: str | None = None
    settore: str | None = None  # CSEA sector estimate (e.g. "Plastica / gomma")
    priorita_fv: str | None = None  # e.g. "Molto alta"
    classe: str | None = None
    tipo_classe: str | None = None
    stato: str | None = None  # e.g. "Confermata" / "In istruttoria"


def normalize_piva(raw: Any) -> str | None:
    """Return an 11-digit P.IVA (zeros preserved) or None if not valid.

    Accepts str or int; strips non-digits; zero-pads a shorter value (a source
    that dropped leading zeros) back to 11; rejects anything longer than 11.
    """
    if raw is None:
        return None
    digits = _DIGITS.sub("", str(raw).strip())
    if not digits or len(digits) > 11:
        return None
    return digits.zfill(11)


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_energivori_rows(
    header: Sequence[Any],
    rows: Iterable[Sequence[Any]],
) -> list[EnergivoroRecord]:
    """Pure core: map rows to records by HEADER NAME (order-independent).

    Skips rows without a valid P.IVA; de-dupes on P.IVA (first wins).
    """
    idx: dict[str, int] = {}
    for i, label in enumerate(header):
        key = str(label or "").strip().lower()
        field = _COLUMN_MAP.get(key)
        if field and field not in idx:
            idx[field] = i

    if "piva" not in idx:
        raise ValueError("energivori: colonna P.IVA non trovata nell'header")

    def cell(row: Sequence[Any], field: str) -> Any:
        i = idx.get(field)
        return row[i] if i is not None and i < len(row) else None

    out: list[EnergivoroRecord] = []
    seen: set[str] = set()
    for row in rows:
        piva = normalize_piva(cell(row, "piva"))
        if piva is None or piva in seen:
            continue
        name = _clean(cell(row, "ragione_sociale"))
        if not name:
            continue
        seen.add(piva)
        out.append(
            EnergivoroRecord(
                piva=piva,
                ragione_sociale=name,
                codice_fiscale=_clean(cell(row, "codice_fiscale")),
                settore=_clean(cell(row, "settore")),
                priorita_fv=_clean(cell(row, "priorita_fv")),
                classe=_clean(cell(row, "classe")),
                tipo_classe=_clean(cell(row, "tipo_classe")),
                stato=_clean(cell(row, "stato")),
            )
        )
    return out


def parse_energivori_xlsx(data: bytes, *, sheet: str | None = None) -> list[EnergivoroRecord]:
    """Read the energivori workbook bytes → records (first row = header)."""
    import io

    import openpyxl  # lazy: only the ingest boundary needs it

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    return parse_energivori_rows(header, it)
